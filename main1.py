'''O objetivo desse codigo é, automatizar planilhas de excel com python,
essa vai ser uma versão simples, que podera ser melhorada futuramente.
Para que o codigo funcione, lembre-se de baixar a biblioteca openpyxl.'''

import openpyxl

#Criando uma planilha
book = openpyxl.Workbook()
#Exibir as planilhas abertas
print(book.sheetnames)
#Criar e selecionar uma pagina
book.create_sheet('Alunos')
alunos_page = book['Alunos']
#Preencher planilha com dados
alunos_page.append(['Alunos','Medias','Aprovações'])
alunos_page.append(['Pedro','9','Aprovado'])
alunos_page.append(['Maria','10','Aprovado'])
alunos_page.append(['Larissa','7','Aprovado'])
alunos_page.append(['Marcos','5','Reprovado'])
alunos_page.append(['Lucas','6','Aprovado'])
alunos_page.append(['Matheus','8','Aprovado'])
#Salvar a planilha
book.save('Planilha de Alunos.xlsx')
